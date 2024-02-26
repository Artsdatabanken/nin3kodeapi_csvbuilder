import os
import requests
import pandas as pd
import sqlite3
# set desired branch to get dbfile from here:
conn = None

def fetchDB(branch:str):
    branch = 'develop' # usually 'develop','test' or 'master'

    url = f"https://raw.githubusercontent.com/Artsdatabanken/nin3-kode-api/{branch}/NiN3.WebApi/nin3kodeapi.db" # Henter nin3 database fra nin3 repo branch:develop
    response = requests.get(url)
    with open("db/nin3kodeapi.db", "wb") as f:
        f.write(response.content)
        print("db-file downloaded to path 'db/nin3kodeapi.db'")

def fetchDBfromLocal(path):
    global conn
    conn = sqlite3.connect(path)
    print(f"db-file connected from local path '{path}'")

def getConn():
    global conn
    if conn is None:
        conn = sqlite3.connect("db/nin3kodeapi.db")
    return conn

def closeConn():
    global conn
    if conn is not None:
        conn.close()
        conn = None

def extend_df_with_edit_logg_columns(df):
    df['Type endring']=None
    df['Dato']=None
    df['Endret av'] = None
    df['Beskrivelse av endring'] = None
    return df
    

def create_so_si_trinn_unihet(dbfromlocal=True, branch='develop'):
    from conf import localdbpath
    import shutil
    if dbfromlocal:
        fetchDBfromLocal(localdbpath)
    else:
        fetchDB(branch)
    query = """select Verdi, Beskrivelse, count(*) as antall_forekomster from Trinn Where MaaleskalaId in 
                           (select id from Maaleskala where MaaleskalaNavn like '%SO' or MaaleskalaNavn like '%SI')
                Group by Verdi, Beskrivelse
                Having count(*)
                order by count(*) desc"""
    so_si_uniket = pd.read_sql_query(query, conn)
    with pd.ExcelWriter(f"ut/trinn_so_si_{timestamp()}.xlsx") as writer:
        so_si_uniket.to_excel(writer, sheet_name="trinn_under_so_si", index=False)
    print(f"Excel data skrevet til 'ut/trinn_so_si.xlsx'")

# Utils
"""    
def autoajustColwith(df, worksheet):
    # adjust the column widths based on the content
    for i, col in enumerate(df.columns):
        if col in worksheet.column_dimensions:
            width = max(df[col].apply(lambda x: len(str(x))).max(), len(col))
            worksheet.column_dimensions[col].width = width
"""


def excel_autoadjust_col(path, target_excel, padding):
    import os
    import openpyxl
    from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
    from openpyxl.utils import get_column_letter

    target_file = os.path.join(path, target_excel)
    wb = openpyxl.load_workbook(target_file)
    sheets = [sheet for sheet in wb.get_sheet_names()]

    for sheet in sheets:
        ws = wb[sheet]
        dim_holder = DimensionHolder(worksheet=ws)

        for col in range(ws.min_column, ws.max_column + 1):
            width = 0
            for row in range(ws.min_row, ws.max_row + 1):
                cell_value = ws.cell(column=col, row=row).value
                if cell_value:
                    cell_len = len(str(cell_value))
                    if cell_len > width:
                        width = cell_len + padding

            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=width)

        ws.column_dimensions = dim_holder

    wb.save(target_file)
    print("Completed adjustments for {}".format(target_excel))

# Main method
def createExcel(dbfromlocal=False, branch='develop', forEdit=True):
    from conf import localdbpath
    import shutil
    if dbfromlocal:
        fetchDBfromLocal(localdbpath)
    else:
        fetchDB(branch)
    conn = getConn()
    df_db_info = db_info_fane()
    df_type = type_fane()
    df_htg = hovedtypegruppe_fane()
    df_ht = hovedtype_fane()
    df_gt = grunntype_fane()
    df_m005 = kle_m005()
    df_m020 = kle_m020()
    df_m050 = kle_m050()
    df_ht_kle005 = ht_kle_m005()
    df_ht_kle020 = ht_kle_m020()
    df_ht_kle050 = ht_kle_m050()
    df_gt_kle005 = gt_kle_m005()
    df_gt_kle020 = gt_kle_m020()
    df_gt_kle050 = gt_kle_m050()
    df_htg_hoek = htg_hoek()
    df_v = variabel_fane()
    df_vn = variabelnavn_fane()
    df_ms = maaleskala_fane()
    df_t = trinn_fane()
    df_vn_ms = variabelnavn_maaleskala()
    df_ht_vt = hovedtype_Variabeltrinn()
    df_gt_vt = grunntype_Variabeltrinn()
    df_htg_konvertering = htg_konvertering()
    df_ht_konvertering = ht_konvertering()
    df_gt_konvertering = gt_konvertering()
    df_vn_konvertering = vn_konvertering()
    df_enums = enums()
    if forEdit:
        df_type = extend_df_with_edit_logg_columns(df_type)
        df_htg = extend_df_with_edit_logg_columns(df_htg)
        df_ht = extend_df_with_edit_logg_columns(df_ht)
        df_gt = extend_df_with_edit_logg_columns(df_gt)
        df_m005 = extend_df_with_edit_logg_columns(df_m005)
        df_m020 = extend_df_with_edit_logg_columns(df_m020)
        df_m050 = extend_df_with_edit_logg_columns(df_m050)
        df_ht_kle005 = extend_df_with_edit_logg_columns(df_ht_kle005)
        df_ht_kle020 = extend_df_with_edit_logg_columns(df_ht_kle020)
        df_ht_kle050 = extend_df_with_edit_logg_columns(df_ht_kle050)
        df_gt_kle005 = extend_df_with_edit_logg_columns(df_gt_kle005)
        df_gt_kle020 = extend_df_with_edit_logg_columns(df_gt_kle020)
        df_gt_kle050 = extend_df_with_edit_logg_columns(df_gt_kle050)
        df_htg_hoek = extend_df_with_edit_logg_columns(df_htg_hoek)
        df_v = extend_df_with_edit_logg_columns(df_v)
        df_vn = extend_df_with_edit_logg_columns(df_vn)
        df_ms = extend_df_with_edit_logg_columns(df_ms)
        df_t = extend_df_with_edit_logg_columns(df_t)
        df_vn_ms = extend_df_with_edit_logg_columns(df_vn_ms)
        df_ht_vt = extend_df_with_edit_logg_columns(df_ht_vt)
        df_gt_vt = extend_df_with_edit_logg_columns(df_gt_vt)
        df_htg_konvertering = extend_df_with_edit_logg_columns(df_htg_konvertering)
        df_ht_konvertering = extend_df_with_edit_logg_columns(df_ht_konvertering)
        df_gt_konvertering = extend_df_with_edit_logg_columns(df_gt_konvertering)
        df_vn_konvertering = extend_df_with_edit_logg_columns(df_vn_konvertering)
        df_enums = extend_df_with_edit_logg_columns(df_enums)
    excelfile = f"ut/nin3_0_{timestamp()}.xlsx"
    with pd.ExcelWriter(excelfile) as writer:
        df_db_info.to_excel(writer, sheet_name="db_info", index=False)
        df_type.to_excel(writer, sheet_name="Type", index=False)
        df_htg.to_excel(writer, sheet_name="Hovedtypegruppe", index=False)
        df_ht.to_excel(writer, sheet_name="Hovedtype", index=False)
        df_gt.to_excel(writer, sheet_name="Grunntype", index=False)
        df_m005.to_excel(writer, sheet_name="KLE_M005", index=False)
        df_m020.to_excel(writer, sheet_name="KLE_M020", index=False)
        df_m050.to_excel(writer, sheet_name="KLE_M050", index=False)
        df_ht_kle005.to_excel(writer, sheet_name="HT_KLE_M005", index=False)
        df_ht_kle020.to_excel(writer, sheet_name="HT_KLE_M020", index=False)
        df_ht_kle050.to_excel(writer, sheet_name="HT_KLE_M050", index=False)
        df_gt_kle005.to_excel(writer, sheet_name="GT_KLE_M005", index=False)
        df_gt_kle020.to_excel(writer, sheet_name="GT_KLE_M020", index=False)
        df_gt_kle050.to_excel(writer, sheet_name="GT_KLE_M050", index=False)
        df_htg_hoek.to_excel(writer, sheet_name="HTG_Hovedøkosystem", index=False)
        df_v.to_excel(writer, sheet_name="Variabel", index=False)
        df_vn.to_excel(writer, sheet_name="Variabelnavn", index=False)
        df_ms.to_excel(writer, sheet_name="Måleskala", index=False)
        df_t.to_excel(writer, sheet_name="Trinn", index=False)
        df_vn_ms.to_excel(writer, sheet_name="Variabelnavn_Måleskala", index=False)
        df_ht_vt.to_excel(writer, sheet_name="Hovedtype_Variabeltrinn", index=False)
        df_gt_vt.to_excel(writer, sheet_name="Grunntype_Variabeltrinn", index=False)
        df_htg_konvertering.to_excel(writer, sheet_name="HTG_Konvertering", index=False)
        df_ht_konvertering.to_excel(writer, sheet_name="HT_Konvertering", index=False)
        df_gt_konvertering.to_excel(writer, sheet_name="GT_Konvertering", index=False)
        df_vn_konvertering.to_excel(writer, sheet_name="VN_Konvertering", index=False)
        df_enums.to_excel(writer, sheet_name="Enums", index=False)
    print(f"\n\nExcel data skrevet til {excelfile}")
    closeConn()
    # add excefile to "new"-folder
    if forEdit:
        destination_file = "ut/new/nin3_0_redigering.xlsx" #file for editing
        shutil.copy(excelfile, destination_file)
        print(f"adjusting column width in: {destination_file}")
        excel_autoadjust_col("ut/new", "nin3_0_redigering.xlsx", 2)
    else:
        destination_file = "ut/api/nin3_0.xlsx"
        shutil.copy(excelfile, destination_file)
        print(f"adjusting column width in: {destination_file}")
        excel_autoadjust_col("ut/api", "nin3_0.xlsx", 2)

    # adjust column width in resulting excels

def timestamp():
    from datetime import datetime
    return datetime.now().strftime('%Y%m%d_%H%M%S')

def db_info_fane(makecsv=False):
    conn = getConn()
    db_infoQ = """SELECT * FROM db_info"""
    df_db_info = pd.read_sql_query(db_infoQ, conn)
    if makecsv:
        df_db_info.to_csv(f"ut/db_info_fane_{timestamp()}.csv", index=False, encoding="utf-8-sig")
        print(f"written to 'ut/db_info_fane_{timestamp()}.csv'")
    return df_db_info

def type_fane(makecsv=False):
    conn = getConn()
    TypefaneQ = """WITH eco AS (
            SELECT Ordinal, Verdi, Beskrivelse 
            FROM Enumoppslag 
            WHERE Enumtype = 'EcosysnivaaEnum'
        ),
        tkat AS (
            SELECT Ordinal, Verdi, Beskrivelse
            FROM Enumoppslag
            WHERE Enumtype = 'TypekategoriEnum'
        ),
        tkat2 AS ( 
            SELECT Ordinal, Verdi, Beskrivelse
            FROM Enumoppslag
            WHERE Enumtype = 'Typekategori2Enum'
        )
        SELECT Langkode, Kode, type.Navn, eco.Verdi AS EcosystnivaaEnum, tkat.Verdi AS Typekategori, tkat2.Verdi AS Typekategori2
        FROM type
        LEFT JOIN eco ON type.Ecosystnivaa = eco.Ordinal
        LEFT JOIN tkat ON type.Typekategori = tkat.Ordinal
        LEFT JOIN tkat2 ON type.Typekategori2 = tkat2.Ordinal"""  # Write your SQLite query
    df_type = pd.read_sql_query(TypefaneQ, conn)  # Execute the query and store the result in a DataFrame
    if makecsv:
        df_type.to_csv(f"ut/type_fane_{timestamp()}.csv", index=False, encoding="utf-8-sig")
        print(f"written to 'ut/type_fane_{timestamp()}.csv'")
    return df_type

def hovedtypegruppe_fane(makecsv=False):
    conn = getConn()
    htgQ = """WITH tkat2 AS ( 
            SELECT Ordinal, Verdi, Beskrivelse
            FROM Enumoppslag
            WHERE Enumtype = 'Typekategori2Enum'
        ), tkat3 AS ( 
            SELECT Ordinal, Verdi, Beskrivelse
            FROM Enumoppslag
            WHERE Enumtype = 'Typekategori3Enum'
        )
        SELECT htg.Langkode, htg.Kode, htg.Delkode, htg.Navn, tkat2.Verdi AS Typekategori2, tkat3.Verdi AS Typekategori3, type.Kode AS Typekode
        FROM hovedtypegruppe htg
        LEFT JOIN tkat2 ON htg.Typekategori2 = tkat2.Ordinal
        LEFT JOIN tkat3 ON htg.Typekategori3 = tkat3.Ordinal
        LEFT JOIN type ON htg.TypeId = type.Id"""
    df_htg = pd.read_sql_query(htgQ, conn)
    if makecsv:
        df_htg.to_csv(f"ut/hovedtypegruppe_fane_{timestamp()}.csv", index=False, encoding="utf-8-sig")
        print(f"written to 'ut/hovedtypegruppe_fane_{timestamp()}.csv'")
    return df_htg


def hovedtype_fane(makecsv=False):
    conn = getConn()
    htQ = """WITH pk AS ( 
        SELECT Ordinal, Verdi, Beskrivelse
        FROM Enumoppslag
        WHERE Enumtype = 'ProsedyrekategoriEnum'
        )
        SELECT ht.Langkode, ht.Kode, ht.Delkode, ht.Navn, pk.Verdi AS Prosedyrekategori, hg.Kode AS Hovedtypegruppekode
        FROM hovedtype ht
        LEFT JOIN hovedtypegruppe hg ON hg.Id = ht.HovedtypegruppeId
        LEFT JOIN pk ON ht.Prosedyrekategori = pk.Ordinal"""
    df_ht = pd.read_sql_query(htQ, conn)
    if makecsv:
        df_ht.to_csv(f"ut/hovedtype_fane_{timestamp()}.csv", index=False, encoding="utf-8-sig")
        print(f"written to 'ut/hovedtype_fane_{timestamp()}.csv'")
    return df_ht


def grunntype_fane(makecsv=False):
    conn = getConn()
    gtQ = """WITH pk AS ( 
    SELECT Ordinal, Verdi, Beskrivelse
    FROM Enumoppslag
    WHERE Enumtype = 'ProsedyrekategoriEnum'
    )
    Select gt.Langkode, gt.Kode, gt.Delkode, gt.Navn, pk.Verdi AS Prosedyrekategori, ht.Kode AS Hovedtypekode
    from grunntype gt
    LEFT JOIN pk ON gt.Prosedyrekategori = pk.Ordinal
    LEFT JOIN hovedtype ht ON gt.HovedtypeId = ht.Id"""
    df_gt = pd.read_sql_query(gtQ, conn)
    if makecsv:
        df_gt.to_csv(f"ut/grunntype_fane_{timestamp()}.csv", index=False, encoding="utf-8-sig")
        print(f"written to 'ut/grunntype_fane_{timestamp()}.csv'")
    return df_gt

def kle_m005(makecsv=False):
    conn = getConn()
    m005Q = """WITH ms AS ( Select Ordinal, Verdi from Enumoppslag where Enumtype like 'MaalestokkEnum' and verdi = 'M005')
    select ke.Langkode, ke.kode, ke.navn, ms.Verdi as Målestokk from Kartleggingsenhet ke
    LEFT JOIN ms ON ms.Ordinal = ke.Maalestokk
    Where ke.Maalestokk = 0
    Order by kode"""
    df_m005 = pd.read_sql_query(m005Q, conn)
    if makecsv:
        df_m005.to_csv(f"ut/m005_fane_{timestamp()}.csv", index=False, encoding="utf-8-sig")
        print(f"written to 'ut/m005_fane_{timestamp()}.csv'")
    return df_m005

def kle_m020(makecsv=False):
    conn = getConn()
    m020Q = """WITH ms AS ( Select Ordinal, Verdi from Enumoppslag where Enumtype like 'MaalestokkEnum' and verdi = 'M020')
    select ke.Langkode, ke.kode, ke.navn, ms.Verdi as Målestokk from Kartleggingsenhet ke
    LEFT JOIN ms ON ms.Ordinal = ke.Maalestokk
    Where ke.Maalestokk = 2
    Order by kode"""
    df_m020 = pd.read_sql_query(m020Q, conn)
    if makecsv:
        df_m020.to_csv(f"ut/m020_fane_{timestamp()}.csv", index=False, encoding="utf-8-sig") 
        print(f"written to 'ut/m020_fane_{timestamp()}.csv'")   
    return df_m020

def kle_m050(makecsv=False):
    conn = getConn()
    m050Q = """WITH ms AS ( Select Ordinal, Verdi from Enumoppslag where Enumtype like 'MaalestokkEnum' and verdi = 'M050')
    select ke.Langkode, ke.kode, ke.navn, ms.Verdi as Målestokk from Kartleggingsenhet ke
    LEFT JOIN ms ON ms.Ordinal = ke.Maalestokk
    Where ke.Maalestokk = 3
    Order by kode"""
    df_m050 = pd.read_sql_query(m050Q, conn)
    if makecsv:
        df_m050.to_csv(f"ut/m050_fane_{timestamp()}.csv", index=False, encoding="utf-8-sig")  
        print(f"written to 'ut/m050_fane_{timestamp()}.csv'")  
    return df_m050

def ht_kle_m005(makecsv=False):
    conn = getConn()
    ht_kle005Q = """With m005 AS (Select Kode, KartleggingsenhetId, HovedtypeId 
              from Kartleggingsenhet ke, Hovedtype_Kartleggingsenhet ht_kle 
              where ke.Maalestokk=0 and ke.Id = ht_kle.KartleggingsenhetId)
              Select ht.Kode as Hovedtype_kode, m005.Kode as M005_kode from Hovedtype ht
              JOIN m005 ON ht.Id = m005.HovedtypeId order by ht.Kode, M005_kode"""
    df_ht_kle005 = pd.read_sql_query(ht_kle005Q, conn)
    if makecsv:
        df_ht_kle005.to_csv(f"ut/hovedtypeKLE_M005_fane_{timestamp()}.csv", index=False, encoding="utf-8-sig") 
        print(f"written to 'ut/hovedtypeKLE_M005_fane_{timestamp()}.csv'")
    return df_ht_kle005

def ht_kle_m020(makecsv=False):
    conn = getConn()
    ht_kle020Q = """with m020 AS (Select Kode, KartleggingsenhetId, HovedtypeId 
                from Kartleggingsenhet ke, Hovedtype_Kartleggingsenhet ht_kle
                where ke.Maalestokk=2 and ke.Id = ht_kle.KartleggingsenhetId)
                Select ht.Kode as Hovedtype_kode, m020.Kode as M020_kode from Hovedtype ht
                JOIN m020 ON ht.Id = m020.HovedtypeId order by ht.Kode, M020_kode"""
    df_ht_kle020 = pd.read_sql_query(ht_kle020Q, conn)
    if makecsv:
        df_ht_kle020.to_csv(f"ut/hovedtypeKLE_M020_fane_{timestamp()}.csv", index=False, encoding="utf-8-sig")   
        print(f"written to 'ut/hovedtypeKLE_M020_fane_{timestamp()}.csv'")
    return df_ht_kle020

def ht_kle_m050(makecsv=False):
    conn = getConn()
    ht_kle050Q = """with m050 AS (Select Kode, KartleggingsenhetId, HovedtypeId
                from Kartleggingsenhet ke, Hovedtype_Kartleggingsenhet ht_kle
                where ke.Maalestokk=3 and ke.Id = ht_kle.KartleggingsenhetId)
                Select ht.Kode as Hovedtype_kode, m050.Kode as M050_kode from Hovedtype ht
                JOIN m050 ON ht.Id = m050.HovedtypeId order by ht.Kode, M050_kode"""
    df_ht_kle050 = pd.read_sql_query(ht_kle050Q, conn)
    if makecsv:
        df_ht_kle050.to_csv(f"ut/hovedtypeKLE_M020_fane_{timestamp()}.csv", index=False, encoding="utf-8-sig") 
        print(f"written to 'ut/hovedtypeKLE_M050_fane_{timestamp()}.csv'")
    return df_ht_kle050

def gt_kle_m005(makecsv=False):
    conn = getConn()
    gt_kle005Q = """With m005 AS 
                (Select Kode, KartleggingsenhetId, GrunntypeId
                from Kartleggingsenhet ke, Kartleggingsenhet_Grunntype gt_kle
                where ke.Maalestokk=0 and ke.Id = gt_kle.KartleggingsenhetId)

                select m005.Kode as M005_kode, gt.Kode as Grunntype_kode 
                from Grunntype gt
                JOIN m005 ON gt.Id = m005.GrunntypeId 
                order by gt.Kode, M005_kode"""
    df_gt_kle005 = pd.read_sql_query(gt_kle005Q, conn)
    if makecsv:
        df_gt_kle005.to_csv(f"ut/grunntypeKLE_M005_fane_{timestamp()}.csv", index=False, encoding="utf-8-sig")
        print(f"written to 'ut/grunntypeKLE_M005_fane_{timestamp()}.csv'")
    return df_gt_kle005

def gt_kle_m020(makecsv=False):
    conn = getConn()
    gt_kle020Q = """With m020 AS (Select Kode, KartleggingsenhetId, GrunntypeId
                from Kartleggingsenhet ke, Kartleggingsenhet_Grunntype gt_kle
                where ke.Maalestokk=2 and ke.Id = gt_kle.KartleggingsenhetId)
                select m020.Kode as M020_kode, gt.Kode as Grunntype_kode from Grunntype gt
                JOIN m020 ON gt.Id = m020.GrunntypeId order by gt.Kode, M020_kode"""
    df_gt_kle020 = pd.read_sql_query(gt_kle020Q, conn)
    if makecsv:
        df_gt_kle020.to_csv(f"ut/grunntypeKLE_M020_fane_{timestamp()}.csv", index=False, encoding="utf-8-sig")
        print(f"written to 'ut/grunntypeKLE_M020_fane_{timestamp()}.csv'")
    return df_gt_kle020

def gt_kle_m050(makecsv=False):
    conn = getConn()
    gt_kle050Q = """With m050 AS (Select Kode, KartleggingsenhetId, GrunntypeId
                from Kartleggingsenhet ke, Kartleggingsenhet_Grunntype gt_kle
                where ke.Maalestokk=3 and ke.Id = gt_kle.KartleggingsenhetId)
                select m050.Kode as M050_kode, gt.Kode as Grunntype_kode from Grunntype gt
                JOIN m050 ON gt.Id = m050.GrunntypeId order by gt.Kode, M050_kode"""
    df_gt_kle050 = pd.read_sql_query(gt_kle050Q, conn)
    if makecsv:
        df_gt_kle050.to_csv(f"ut/grunntypeKLE_M050_fane_{timestamp()}.csv", index=False, encoding="utf-8-sig")
        print(f"written to 'ut/grunntypeKLE_M050_fane_{timestamp()}.csv'")
    return df_gt_kle050

def htg_hoek(makecsv=False):
    conn = getConn()
    htg_hoekQ = """with hs AS (Select Ordinal, Verdi from Enumoppslag where Enumtype like 'HovedoekosystemEnum')
    select Kode, hs.Verdi from Hovedtypegruppe htg
    JOIN Hovedtypegruppe_Hovedoekosystem htg_ho ON htg.Id = htg_ho.HovedtypegruppeId
    JOIN hs ON hs.Ordinal = htg_ho.HovedoekosystemEnum
    Order by Kode"""
    df_htg_hoek = pd.read_sql_query(htg_hoekQ, conn)
    if makecsv:
        df_htg_hoek.to_csv(f"ut/htg_hoek_fane_{timestamp()}.csv", index=False, encoding="utf-8-sig")
        print(f"written to 'ut/htg_hoek_fane_{timestamp()}.csv'")
    return df_htg_hoek

def variabel_fane(makecsv=False):
    conn = getConn()
    v_Q = """WITH eco AS ( 
        SELECT Ordinal, Verdi, Beskrivelse
        FROM Enumoppslag
        WHERE Enumtype = 'EcosysnivaaEnum'
    ),
    vk AS (
        SELECT Ordinal, Verdi, Beskrivelse
        FROM Enumoppslag
        WHERE Enumtype = 'VariabelkategoriEnum' 
    )
    SELECT v.Langkode, v.Kode, v.Navn, eco.Verdi AS Ecosysnivaa, vk.Verdi AS Variabelkategori
    FROM Variabel v
    LEFT JOIN eco ON v.Ecosystnivaa = eco.Ordinal
    LEFT JOIN vk ON v.Variabelkategori = vk.Ordinal"""
    df_v = pd.read_sql_query(v_Q, conn)
    if makecsv:
        df_v.to_csv(f"ut/variabel_fane_{timestamp()}.csv", index=False, encoding="utf-8-sig")
        print(f"written to 'ut/variabel_fane_{timestamp()}.csv'")
    return df_v

def variabelnavn_fane(makecsv=False):
    conn = getConn()
    vn_Q = """WITH vk2 AS (
        SELECT Ordinal, Verdi, Beskrivelse
        FROM Enumoppslag
        WHERE Enumtype = 'Variabelkategori2Enum'
    ),
    vt AS (
        SELECT Ordinal, Verdi, Beskrivelse
        FROM Enumoppslag
        WHERE Enumtype = 'VariabeltypeEnum'
    ),
    vg AS (
        SELECT Ordinal, Verdi, Beskrivelse
        FROM Enumoppslag
        WHERE Enumtype = 'VariabelgruppeEnum'
    ),
    v AS(select Id, Kode From Variabel)   
    SELECT vn.Langkode, vn.Kode, vn.Navn, vk2.Verdi AS Variabelkategori2, vt.Verdi AS Variabeltype, vg.Verdi AS Variabelgruppe, v.Kode AS Variabelkode
    FROM Variabelnavn vn
    LEFT JOIN vk2 ON vn.Variabelkategori2 = vk2.Ordinal
    LEFT JOIN vt ON vn.Variabeltype = vt.Ordinal
    LEFT JOIN vg ON vn.Variabelgruppe = vg.Ordinal
    LEFT JOIN v ON vn.VariabelId = v.Id"""
    df_vn = pd.read_sql_query(vn_Q, conn)
    if makecsv:
        df_vn.to_csv(f"ut/variabelnavn_fane_{timestamp()}.csv", index=False, encoding="utf-8-sig")
        print(f"written to 'ut/variabelnavn_fane_{timestamp()}.csv'")
    return df_vn

def maaleskala_fane(makecsv=False):
    conn = getConn()
    ms_Q = """WITH ee AS (
        SELECT Ordinal, Verdi, Beskrivelse
        FROM Enumoppslag
        WHERE Enumtype = 'EnhetEnum'
    ),
    me AS (
        SELECT Ordinal, Verdi, Beskrivelse
        FROM Enumoppslag
        WHERE Enumtype = 'MaaleskalatypeEnum'
    )
    SELECT m.MaaleskalaNavn, ee.Verdi AS Enhet, me.Verdi AS Maaleskalatype
    FROM Maaleskala m
    LEFT JOIN ee ON m.EnhetEnum = ee.Ordinal
    LEFT JOIN me ON m.MaaleskalatypeEnum = me.Ordinal"""
    df_ms = pd.read_sql_query(ms_Q, conn)
    if makecsv:
        df_ms.to_csv(f"ut/maaleskala_fane_{timestamp()}.csv", index=False, encoding="utf-8-sig")
        print(f"written to 'ut/maaleskala_fane_{timestamp()}.csv'")
    return df_ms

def trinn_fane(makecsv=False):
    conn = getConn()
    t_Q = """
    SELECT t.Verdi, t.Beskrivelse, ms.MaaleskalaNavn
    FROM Trinn t
    LEFT JOIN Maaleskala ms ON t.MaaleskalaId = ms.Id"""
    df_t = pd.read_sql_query(t_Q, conn)
    if makecsv:
        df_t.to_csv(f"ut/trinn_fane_{timestamp()}.csv", index=False, encoding="utf-8-sig")
        print(f"written to 'ut/trinn_fane_{timestamp()}.csv'")
    return df_t

def variabelnavn_maaleskala(makecsv=False):
    conn = getConn()
    vn_msQ = """  
    select vn.Kode AS VariabelnavnKode, vn.navn AS Variabelnavn, ms.MaaleskalaNavn
    FROM VariabelnavnMaaleskala vnm
    LEFT JOIN Maaleskala ms ON vnm.MaaleskalaId = ms.id
    LEFT JOIN Variabelnavn vn ON vnm.VariabelnavnId = vn.id
    Order by VariabelnavnKode"""
    df_vn_ms = pd.read_sql_query(vn_msQ, conn)
    if makecsv:
        df_vn_ms.to_csv(f"ut/variabelnavn_maaleskala_fane_{timestamp()}.csv", index=False, encoding="utf-8-sig")
        print(f"written to 'ut/variabelnavn_maaleskala_fane_{timestamp()}.csv'")
    return df_vn_ms

def hovedtype_Variabeltrinn(makecsv=False):
    conn = getConn()
    ht_vtQ = """Select ht.Kode as HovedtypeKode, vn.Kode as VariabelnavnKode, ms.MaaleskalaNavn, t.Verdi as TrinnVerdi
    from HovedtypeVariabeltrinn hvt
    LEFT JOIN Variabelnavn vn ON hvt.VariabelnavnId = vn.Id
    LEFT JOIN Hovedtype ht ON hvt.HovedtypeId = ht.Id
    LEFT JOIN Maaleskala ms ON hvt.MaaleskalaId = ms.Id
    LEFT JOIN Trinn t ON hvt.TrinnId = t.Id
    ORDER BY ht.Kode, VariabelnavnKode, MaaleskalaNavn, TrinnVerdi"""
    df_ht_vt = pd.read_sql_query(ht_vtQ, conn)
    if makecsv:
        df_ht_vt.to_csv(f"ut/hovedtype_variabeltrinn_fane_{timestamp()}.csv", index=False, encoding="utf-8-sig")
        print(f"written to 'ut/hovedtype_variabeltrinn_fane_{timestamp()}.csv'")
    return df_ht_vt

def grunntype_Variabeltrinn(makecsv=False):
    conn = getConn()
    gt_vtQ = """Select gt.Kode as GrunntypeKode, vn.Kode as VariabelnavnKode, ms.MaaleskalaNavn, t.Verdi as TrinnVerdi
    from GrunntypeVariabeltrinn gvt
    LEFT JOIN Variabelnavn vn ON gvt.VariabelnavnId = vn.Id
    LEFT JOIN Grunntype gt ON gvt.GrunntypeId = gt.Id
    LEFT JOIN Maaleskala ms ON gvt.MaaleskalaId = ms.Id
    LEFT JOIN Trinn t ON gvt.TrinnId = t.Id
    ORDER BY gt.Kode, VariabelnavnKode, MaaleskalaNavn, TrinnVerdi"""
    df_gt_vt = pd.read_sql_query(gt_vtQ, conn)
    if makecsv:
        df_gt_vt.to_csv(f"ut/grunntype_variabeltrinn_fane_{timestamp()}.csv", index=False, encoding="utf-8-sig")
        print(f"written to 'ut/grunntype_variabeltrinn_fane_{timestamp()}.csv'")
    return df_gt_vt

def htg_konvertering(makecsv=False):
    conn = getConn()
    htg_konverteringQ = """with t as (
    SELECT Ordinal
    FROM Enumoppslag
    WHERE Enumtype = 'KlasseEnum' and Beskrivelse = 'Hovedtypegruppe')
    Select htg.Langkode, k.Kode, k.ForrigeKode, k.FoelsomhetsPresisjon as FP, k.Spesifiseringsevne as SP, k.Url
    FROM Konvertering k
    LEFT JOIN Hovedtypegruppe htg on k.Kode = htg.Kode
    JOIN t ON t.Ordinal = k.Klasse"""
    df_htg_konvertering = pd.read_sql_query(htg_konverteringQ, conn)
    if makecsv:
        df_htg_konvertering.to_csv(f"ut/htg_konvertering_fane_{timestamp()}.csv", index=False, encoding="utf-8-sig")
        print(f"written to 'ut/htg_konvertering_fane_{timestamp()}.csv'")
    return df_htg_konvertering

def ht_konvertering(makecsv=False):
    conn = getConn()
    ht_konverteringQ = """with t as (SELECT Ordinal
        FROM Enumoppslag
        WHERE Enumtype = 'KlasseEnum' and Beskrivelse = 'Hovedtype')
        Select ht.Langkode, k.Kode, k.ForrigeKode, k.FoelsomhetsPresisjon as FP, k.Spesifiseringsevne as SP, k.Url
        FROM Konvertering k
        LEFT JOIN Hovedtype ht on k.Kode = ht.Kode
        JOIN t ON t.Ordinal = k.Klasse"""
    df_ht_konvertering = pd.read_sql_query(ht_konverteringQ, conn)
    if makecsv:
        df_ht_konvertering.to_csv(f"ut/ht_konvertering_fane_{timestamp()}.csv", index=False, encoding="utf-8-sig")
        print(f"written to 'ut/ht_konvertering_fane_{timestamp()}.csv'")
    return df_ht_konvertering

def gt_konvertering(makecsv=False):
    conn = getConn()
    gt_konverteringQ = """with t as (SELECT Ordinal
        FROM Enumoppslag
        WHERE Enumtype = 'KlasseEnum' and Beskrivelse = 'Grunntype')
        Select gt.Langkode, k.Kode, k.ForrigeKode, k.FoelsomhetsPresisjon as FP, k.Spesifiseringsevne as SP, k.Url
        FROM Konvertering k
        LEFT JOIN Grunntype gt on k.Kode = gt.Kode
        JOIN t ON t.Ordinal = k.Klasse"""
    df_gt_konvertering = pd.read_sql_query(gt_konverteringQ, conn)
    if makecsv:
        df_gt_konvertering.to_csv(f"ut/gt_konvertering_fane_{timestamp()}.csv", index=False, encoding="utf-8-sig")
        print(f"written to 'ut/gt_konvertering_fane_{timestamp()}.csv'")
    return df_gt_konvertering

def vn_konvertering(makecsv=False):
    conn = getConn()
    vn_konverteringQ = """with t as (SELECT Ordinal
        FROM Enumoppslag
        WHERE Enumtype = 'KlasseEnum' and Beskrivelse = 'Variabelnavn')
        Select vn.Langkode, k.Kode, k.ForrigeKode, k.FoelsomhetsPresisjon as FP, k.Spesifiseringsevne as SP, k.Url
        FROM Konvertering k
        LEFT JOIN Variabelnavn vn on k.Kode = vn.Kode
        JOIN t ON t.Ordinal = k.Klasse"""
    df_vn_konvertering = pd.read_sql_query(vn_konverteringQ, conn)
    if makecsv:
        df_vn_konvertering.to_csv(f"ut/vn_konvertering_fane_{timestamp()}.csv", index=False, encoding="utf-8-sig")
        print(f"written to 'ut/vn_konvertering_fane_{timestamp()}.csv'")
    return df_vn_konvertering

def enums(makecsv=False):
    conn = getConn()
    enumsQ = """select Enumtype, Verdi, Beskrivelse 
                from enumoppslag order by Enumtype, Ordinal"""
    df_enums = pd.read_sql_query(enumsQ, conn)
    if makecsv:
        df_enums.to_csv(f"ut/enums_{timestamp()}.csv", index=False, encoding="utf-8-sig")
        print(f"written to 'ut/enums_{timestamp()}.csv'")
    return df_enums

# grunntype_maalestokk_rapport.xlsx
def createGrunntypeMaalestokkVariabelOversikt():
    conn = getConn()
    df_db_info = db_info_fane()
    df_gt_m005 = grunntype_m005()
    df_gt_m020 = grunntype_m020()
    variabel_fane = variabelnavn_fane()
    excelfile = f"ut/nin3_0_GT_Maalestokk_VN_{timestamp()}.xlsx"
    with pd.ExcelWriter(excelfile) as writer:
        df_db_info.to_excel(writer, sheet_name="db_info", index=False)
        df_gt_m005.to_excel(writer, sheet_name="Grunntype_M005_VN", index=False)
        df_gt_m020.to_excel(writer, sheet_name="Grunntype_M020_VN", index=False)
        variabel_fane.to_excel(writer, sheet_name="Variabelnavn", index=False)
    print(f"Excel data skrevet til {excelfile}")
    closeConn()
    
def grunntype_m005(makecsv=False):
    conn = getConn()
    gt_m005Q = """WITH m005 AS (
                SELECT ke.Kode, ke.Navn as M005_Navn, KartleggingsenhetId, GrunntypeId
                FROM Kartleggingsenhet ke, Kartleggingsenhet_Grunntype gt_kle
                WHERE ke.Maalestokk = 0 
                AND ke.Id = gt_kle.KartleggingsenhetId
            ),
            gtvt_vn AS (
                SELECT GrunntypeId, vn.Kode
                FROM grunntypeVariabeltrinn gtv, Variabelnavn vn
                WHERE gtv.VariabelnavnId = vn.Id
            )

            SELECT gt.Kode AS Grunntype_kode, gt.Navn AS Grunntypenavn, m005.Kode AS M005_kode, m005.M005_Navn, gtvt_vn.Kode AS Variabelnavn_kode
            FROM Grunntype gt
            JOIN m005 ON gt.Id = m005.GrunntypeId
            LEFT JOIN gtvt_vn ON gt.Id = gtvt_vn.GrunntypeId 
            ORDER BY gt.Kode, M005_kode"""
    df_gt_m005 = pd.read_sql_query(gt_m005Q, conn)
    return df_gt_m005

def grunntype_m020(makecsv=False):
    conn = getConn()
    gt_m020Q = """WITH m020 AS (
            SELECT ke.Kode, ke.Navn as M020_Navn, KartleggingsenhetId, GrunntypeId
            FROM Kartleggingsenhet ke, Kartleggingsenhet_Grunntype gt_kle
            WHERE ke.Maalestokk = 2 
            AND ke.Id = gt_kle.KartleggingsenhetId
        ),
        gtvt_vn AS (
            SELECT GrunntypeId, vn.Kode
            FROM grunntypeVariabeltrinn gtv, Variabelnavn vn
            WHERE gtv.VariabelnavnId = vn.Id
        )

        SELECT gt.Kode AS Grunntype_kode, gt.Navn AS Grunntypenavn, m020.Kode AS M020_kode, m020.M020_Navn, gtvt_vn.Kode AS Variabelnavn_kode
        FROM Grunntype gt
        JOIN m020 ON gt.Id = m020.GrunntypeId
        LEFT JOIN gtvt_vn ON gt.Id = gtvt_vn.GrunntypeId 
        ORDER BY gt.Kode, M020_kode"""
    df_gt_m020 = pd.read_sql_query(gt_m020Q, conn)
    return df_gt_m020

